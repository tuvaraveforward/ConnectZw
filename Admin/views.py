from django import forms
from .models import Dealer

# DealerEditForm must be defined after Dealer is imported
class DealerEditForm(forms.ModelForm):
    class Meta:
        model = Dealer
        fields = ['username', 'firstname', 'lastname', 'phone_number', 'email_address', 'category', 'is_active']




def edit_user(request, user_id):
    try:
        dealer = Dealer.objects.get(id=user_id)
    except Dealer.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('locations')

    if request.method == 'POST':
        form = DealerEditForm(request.POST, instance=dealer)
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully!')
            return redirect('locations')
    else:
        form = DealerEditForm(instance=dealer)

    return render(request, 'admin/edit_dealer.html', {'form': form, 'dealer': dealer, 'user_id': user_id})
    

def delete_user(request, user_id):
    try:
        dealer = Dealer.objects.get(id=user_id)
    except Dealer.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('locations')

    if request.method == 'POST':
        dealer.delete()
        messages.success(request, 'User deleted successfully!')
        return redirect('locations')

    return render(request, 'admin/delete_user.html', {'dealer': dealer, 'user_id': user_id})


def edit_admin(request, admin_id):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    try:
        admin_obj = Admin.objects.get(id=admin_id)
    except Admin.DoesNotExist:
        messages.error(request, 'Admin user not found.')
        return redirect('settings')

    if request.method == 'POST':
        form = AdminEditForm(request.POST, instance=admin_obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Admin updated successfully!')
            return redirect('settings')
    else:
        form = AdminEditForm(instance=admin_obj)

    return render(request, 'admin/edit_admin.html', {'form': form, 'admin_obj': admin_obj, 'admin_id': admin_id})


def delete_admin(request, admin_id):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    try:
        admin_obj = Admin.objects.get(id=admin_id)
    except Admin.DoesNotExist:
        messages.error(request, 'Admin user not found.')
        return redirect('settings')

    if request.method == 'POST':
        admin_obj.delete()
        messages.success(request, 'Admin deleted successfully!')
        return redirect('settings')

    return render(request, 'admin/delete_admin.html', {'admin_obj': admin_obj, 'admin_id': admin_id})
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from django.forms import modelformset_factory
from .models import Admin, Transaction, Feedback as AdminFeedback, LoginAttempt
from Client.models import Feedback as ClientFeedback
from .forms import AdminSignupForm, LocationRegistrationForm, UserRegistrationForm,PosRegistrationForm, ClientEditForm, TransactionForm, DealerForm, ReportForm
from Location.models import Location, User
from Dealer.models import Dealer
from Pos.models import Pos
from Client.models import Client
from .utils import send_account_notification

# Admin edit form (defined after Admin model import)
class AdminEditForm(forms.ModelForm):
    class Meta:
        model = Admin
        fields = ['username', 'firstname', 'lastname', 'phone_number', 'email_address']

# Create your views here.
def main_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        attempt, created = LoginAttempt.objects.get_or_create(username=username)
        if attempt.is_blocked:
            messages.error(request, 'Account blocked due to 3 failed login attempts. Please contact support.')
            return render(request, 'login.html')

        # 1. Check Admin
        try:
            admin_user = Admin.objects.get(username=username)
            if check_password(password, admin_user.password):
                attempt.failed_attempts = 0
                attempt.save()
                request.session['admin_logged_in'] = True
                request.session['admin_username'] = admin_user.username
                return redirect('admin_login')
        except Admin.DoesNotExist:
            pass

        # 2. Check Client
        try:
            client_user = Client.objects.get(username=username)
            if check_password(password, client_user.password):
                attempt.failed_attempts = 0
                attempt.save()
                request.session['client_logged_in'] = True
                request.session['client_username'] = client_user.username
                return redirect('client_login')
        except Client.DoesNotExist:
            pass

        # 3. Check Dealer
        try:
            dealer_user = Dealer.objects.get(username=username)
            if dealer_user.check_password(password):
                attempt.failed_attempts = 0
                attempt.save()
                request.session['dealer_logged_in'] = True
                request.session['dealer_id'] = dealer_user.id
                request.session['dealer_username'] = dealer_user.username
                return redirect('dealer_login')
        except Dealer.DoesNotExist:
            pass

        # 4. Check Location
        try:
            location_user = Location.objects.get(username=username)
            if location_user.check_password(password):
                attempt.failed_attempts = 0
                attempt.save()
                request.session['location_logged_in'] = True
                request.session['location_id'] = location_user.id
                request.session['location_username'] = location_user.username
                return redirect('location_login')
        except Location.DoesNotExist:
            pass

        # 5. Check Pos (Location User/Staff)
        try:
            pos_user = User.objects.get(username=username)
            if pos_user.check_password(password):
                attempt.failed_attempts = 0
                attempt.save()
                request.session['pos_logged_in'] = True
                request.session['pos_user_id'] = pos_user.id
                request.session['pos_username'] = pos_user.username
                return redirect('pos_login')
        except User.DoesNotExist:
            pass
            
        attempt.failed_attempts += 1
        if attempt.failed_attempts >= 3:
            attempt.is_blocked = True
            messages.error(request, 'Account blocked due to 3 failed login attempts. Please contact support.')
        else:
            messages.error(request, 'Invalid credentials')
        attempt.save()
        
    return render(request, 'login.html')

def admin_login(request):
    if request.method == 'POST':
        account_name = request.POST.get('username')
        password = request.POST.get('password')
        
        attempt, created = LoginAttempt.objects.get_or_create(username=account_name)
        if attempt.is_blocked:
            messages.error(request, 'Account blocked due to 3 failed login attempts. Please contact support.')
            return render(request, 'admin/admin_login.html')

        try:
            admin_user = Admin.objects.get(username=account_name)
            if check_password(password, admin_user.password):
                attempt.failed_attempts = 0
                attempt.save()
                request.session['admin_logged_in'] = True
                request.session['admin_username'] = admin_user.username
                return redirect('admin_dashboard')
            else:
                attempt.failed_attempts += 1
                if attempt.failed_attempts >= 3:
                    attempt.is_blocked = True
                    messages.error(request, 'Account blocked due to 3 failed login attempts. Please contact support.')
                else:
                    messages.error(request, 'Invalid credentials')
                attempt.save()
        except Admin.DoesNotExist:
            attempt.failed_attempts += 1
            if attempt.failed_attempts >= 3:
                attempt.is_blocked = True
                messages.error(request, 'Account blocked due to 3 failed login attempts. Please contact support.')
            else:
                messages.error(request, 'Invalid credentials')
            attempt.save()
    return render(request, 'admin/admin_login.html')

def admin_dashboard(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')
    username = request.session.get('admin_username')
    return render(request, 'admin/admin_dashboard.html', {'username': username})


def admin_feedbacks(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    # filter_type: 'feedback' or 'complaint' or None for all
    filter_type = request.GET.get('type')

    admin_feedback_qs = AdminFeedback.objects.all()
    client_feedback_qs = ClientFeedback.objects.all()

    if filter_type:
        admin_feedback_qs = admin_feedback_qs.filter(feedback_type=filter_type)
        client_feedback_qs = client_feedback_qs.filter(feedback_type=filter_type)

    combined = []
    for f in admin_feedback_qs:
        combined.append({
            'feedback_type': f.feedback_type,
            'subject': f.subject,
            'message': f.message,
            'source': 'Location',
            'source_name': getattr(f.location, 'location_name', '') if hasattr(f, 'location') else '',
            'created_at': getattr(f, 'created_at', None),
        })

    for f in client_feedback_qs:
        combined.append({
            'feedback_type': f.feedback_type,
            'subject': f.subject,
            'message': f.message,
            'source': 'Client',
            'source_name': getattr(f.client, 'username', '') if hasattr(f, 'client') else '',
            'created_at': getattr(f, 'created_at', None),
        })

    # sort combined by created_at desc (None at end)
    combined.sort(key=lambda x: x['created_at'] or 0, reverse=True)

    return render(request, 'admin/feedback_list.html', {
        'feedbacks': combined,
        'filter_type': filter_type or 'all',
    })

def admin_signup(request):
    # Signup Disabled
    return redirect('admin_login')
    # if request.method == 'POST':
    #     form = AdminSignupForm(request.POST)
    #     if form.is_valid():
    #         admin = form.save(commit=False)
    #         admin.password = make_password(form.cleaned_data['password'])
    #         admin.save()
    #         messages.success(request, 'Account created successfully. Please log in.')
    #         return redirect('admin_login')
    # else:
    #     form = AdminSignupForm()
    # return render(request, 'admin/admin_signup.html', {'form': form})

def locations(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    from datetime import datetime, timedelta
    from django.db.models import Q
    from django.utils import timezone

    # Get filter parameters from GET request
    filter_type = request.GET.get('filter', 'all')
    show_category = request.GET.get('category', 'off') == 'on'
    search_query = request.GET.get('search', '')

    # Initialize filter flags
    show_new = False
    show_active = False
    show_blocked = False

    # Set flags based on filter type
    if filter_type == 'new':
        show_new = True
    elif filter_type == 'active':
        show_active = True
    elif filter_type == 'blocked':
        show_blocked = True
    else:  # 'all' or default
        show_new = show_active = show_blocked = True

    locations_list = Location.objects.all()

    # Apply filters only if not showing all
    if filter_type != 'all':
        filter_conditions = Q()
        if show_new:
            # Show locations added within the last 24 hours
            twenty_four_hours_ago = timezone.now() - timedelta(hours=24)
            filter_conditions |= Q(created_at__gte=twenty_four_hours_ago)
        if show_active:
            filter_conditions |= Q(is_active=True)
        if show_blocked:
            filter_conditions |= Q(is_active=False)

        locations_list = locations_list.filter(filter_conditions)

    # Apply search filter
    if search_query:
        search_conditions = Q()
        search_conditions |= Q(location_name__icontains=search_query)
        search_conditions |= Q(username__icontains=search_query)
        search_conditions |= Q(email_address__icontains=search_query)
        search_conditions |= Q(city__icontains=search_query)
        search_conditions |= Q(province__icontains=search_query)
        locations_list = locations_list.filter(search_conditions)

    return render(request, 'admin/locations.html', {
        'locations': locations_list,
        'show_new': show_new,
        'show_active': show_active,
        'show_blocked': show_blocked,
        'show_category': show_category,
        'search_query': search_query,
    })

def register_location(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    if request.method == 'POST':
        form = LocationRegistrationForm(request.POST)
        if form.is_valid():
            try:
                location = form.save()
                # Send email notification
                password = form.cleaned_data.get('password')
                send_account_notification(location, password, 'location', request)
                
                messages.success(request, 'Location registered successfully!')
                return redirect('locations')
            except Exception as e:
                messages.error(request, f'Error registering location: {str(e)}')
        else:
            # Add form errors to messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = LocationRegistrationForm()

    return render(request, 'admin/locations.html', {'form': form})
def edit_location(request, location_id):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        messages.error(request, 'Location not found.')
        return redirect('locations')

    # Fetch only dealers and POS for the current location
    # Get users associated with the POS devices at this location
    user_ids = Pos.objects.filter(location=location).values_list('user_id', flat=True).distinct()
    users = User.objects.filter(id__in=user_ids)
    pos_devices = Pos.objects.filter(location=location)

    if request.method == 'POST':
        form = LocationRegistrationForm(request.POST, instance=location)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Location updated successfully!')
                return redirect('locations')
            except Exception as e:
                messages.error(request, f'Error updating location: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = LocationRegistrationForm(instance=location)

    return render(request, 'admin/edit_location.html', {
        'form': form,
        'location': location,
        'users': users,
        'pos_devices': pos_devices,
    })

def delete_location(request, location_id):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')
    
    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        messages.error(request, 'Location not found.')
        return redirect('locations')
    
    if request.method == 'POST':
        location.delete()
        messages.success(request, f'Location "{location.location_name}" deleted successfully!')
        return redirect('locations')
    
    return redirect('edit_location', location_id=location_id)


def add_user(request, location_id):

    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        messages.error(request, 'Location not found.')
        return redirect('locations')

    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                # Create POS for the user at this location
                pos = Pos.objects.create(
                    pos_name=f"{user.username}_POS",
                    location=location,
                    user=user,
                    is_active=True
                )
                
                # Send email to User
                send_account_notification(user, form.cleaned_data['password'], 'user', request)
                
                # Send email about POS (to User's email as per utils logic)
                send_account_notification(pos, pos.pos_code, 'pos', request)
                messages.success(request, f'User "{user.username}" and POS "{pos.pos_name}" created successfully!')
                return redirect('locations')
            except Exception as e:
                messages.error(request, f'Error creating user and POS: {str(e)}')
        else:
            # Add form errors to messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = UserRegistrationForm()

    return render(request, 'admin/add_user_pos.html', {'form': form, 'location': location})

def add_pos(request, location_id):
    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        messages.error(request, 'Location not found.')
        return redirect('locations')

    if request.method == 'POST':
        form = PosRegistrationForm(request.POST)
        if form.is_valid():
            pos = form.save(commit=False)
            pos.location = location
            pos.save()
            
            # Send email to Location (or User if assigned)
            send_account_notification(pos, pos.pos_code, 'pos', request)
            
            messages.success(request, f'POS "{pos.pos_name}" created successfully for location "{location.location_name}"!')
            return redirect('edit_location', location_id=location.id)
        else:
            # Add form errors to messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        # Exclude the location field from the form as it's determined by the URL
        form = PosRegistrationForm(initial={'location': location})
        form.fields['location'].widget = forms.HiddenInput()

    return render(request, 'admin/add_user_pos.html', {'form': form, 'location': location})

def accounts(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    from datetime import datetime, timedelta
    from django.db.models import Q
    from django.utils import timezone

    # Get filter parameters from GET request
    filter_type = request.GET.get('filter', 'all')
    search_query = request.GET.get('search', '')

    # Initialize filter flags
    show_new = False
    show_active = False
    show_blocked = False

    # Set flags based on filter type
    if filter_type == 'new':
        show_new = True
    elif filter_type == 'active':
        show_active = True
    elif filter_type == 'blocked':
        show_blocked = True
    else:  # 'all' or default
        show_new = show_active = show_blocked = True

    clients = Client.objects.all()

    # Apply filters only if not showing all
    if filter_type != 'all':
        filter_conditions = Q()
        if show_new:
            # Show accounts created within the last 24 hours
            twenty_four_hours_ago = timezone.now() - timedelta(hours=24)
            filter_conditions |= Q(created_at__gte=twenty_four_hours_ago)
        if show_active:
            filter_conditions |= Q(is_active=True)
        if show_blocked:
            filter_conditions |= Q(is_active=False)

        clients = clients.filter(filter_conditions)

    # Apply search filter
    if search_query:
        search_conditions = Q()
        search_conditions |= Q(username__icontains=search_query)
        search_conditions |= Q(firstname__icontains=search_query)
        search_conditions |= Q(lastname__icontains=search_query)
        search_conditions |= Q(email_address__icontains=search_query)
        search_conditions |= Q(phone_number__icontains=search_query)
        clients = clients.filter(search_conditions)

    return render(request, 'admin/accounts.html', {
        'clients': clients,
        'show_new': show_new,
        'show_active': show_active,
        'show_blocked': show_blocked,
        'search_query': search_query,
    })

def edit_client(request, client_id):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    try:
        client = Client.objects.get(id=client_id)
    except Client.DoesNotExist:
        messages.error(request, 'Client not found.')
        return redirect('accounts')

    if request.method == 'POST':
        form = ClientEditForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, 'Client updated successfully!')
            return redirect('accounts')
    else:
        form = ClientEditForm(instance=client)

    return render(request, 'admin/edit_client.html', {'form': form, 'client': client, 'client_id': client_id})

def delete_client(request, client_id):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')
    
    try:
        client = Client.objects.get(id=client_id)
    except Client.DoesNotExist:
        messages.error(request, 'Client not found.')
        return redirect('accounts')
    
    if request.method == 'POST':
        client_username = client.username
        client.delete()
        messages.success(request, f'Client "{client_username}" deleted successfully!')
        return redirect('accounts')
    
    return redirect('edit_client', client_id=client_id)

def sales(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    transactions = Transaction.objects.all().select_related('client', 'dealer', 'location', 'pos')

    # Filter parameters
    transaction_type_filter = request.GET.get('type', '')
    client_filter = request.GET.get('client', '')
    dealer_filter = request.GET.get('dealer', '')
    location_filter = request.GET.get('location', '')

    if transaction_type_filter:
        if transaction_type_filter == 'redemption':
             from django.db.models import Q
             # Show true redemptions OR purchases that have been redeemed
             transactions = transactions.filter(Q(transaction_type='redemption') | Q(is_redeemed=True))
        else:
             transactions = transactions.filter(transaction_type=transaction_type_filter)
    if client_filter:
        transactions = transactions.filter(client__username__icontains=client_filter)
    if dealer_filter:
        transactions = transactions.filter(dealer__username__icontains=dealer_filter)
    if location_filter:
        from django.db.models import Q
        transactions = transactions.filter(
            Q(location__location_name__icontains=location_filter) |
            Q(redemption_pos__location__location_name__icontains=location_filter)
        )

    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transaction recorded successfully!')
            return redirect('sales')
    else:
        form = TransactionForm()

    context = {
        'transactions': transactions,
        'form': form,
        'transaction_type_filter': transaction_type_filter,
        'client_filter': client_filter,
        'dealer_filter': dealer_filter,
        'location_filter': location_filter,
    }
    return render(request, 'admin/sales.html', context)

def settings(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    DealerFormSet = modelformset_factory(Dealer, fields=['category'], extra=0, can_delete=False)

    if request.method == 'POST':
        if 'add_dealer' in request.POST:
            dealer_form = DealerForm(request.POST)
            if dealer_form.is_valid():
                dealer = dealer_form.save(commit=False)
                dealer.password = make_password(dealer_form.cleaned_data['password'])
                dealer.save()
                
                # Send email notification
                password = dealer_form.cleaned_data.get('password')
                send_account_notification(dealer, password, 'dealer', request)

                messages.success(request, f'Dealer "{dealer.username}" added successfully!')
                return redirect('settings')
            else:
                for field, errors in dealer_form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        elif 'add_admin' in request.POST:
            admin_form = AdminSignupForm(request.POST)
            if admin_form.is_valid():
                admin = admin_form.save(commit=False)
                admin.password = make_password(admin_form.cleaned_data['password'])
                admin.save()
                messages.success(request, f'Admin "{admin.username}" added successfully!')
                return redirect('settings')
            else:
                for field, errors in admin_form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        elif 'update_categories' in request.POST:
            formset = DealerFormSet(request.POST)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Dealer categories updated successfully!')
                return redirect('settings')
            else:
                messages.error(request, 'Error updating categories.')

    dealer_form = DealerForm()
    admin_form = AdminSignupForm()
    formset = DealerFormSet(queryset=Dealer.objects.all())

    # Get all admins and dealers for display
    admins = Admin.objects.all()
    dealers = Dealer.objects.all()

    context = {
        'dealer_form': dealer_form,
        'admin_form': admin_form,
        'formset': formset,
        'admins': admins,
        'dealers': dealers,
    }
    return render(request, 'admin/settings.html', context)

def reports(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            # Get filter parameters
            dealer = form.cleaned_data.get('dealer')
            client = form.cleaned_data.get('client')
            location = form.cleaned_data.get('location')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            email = form.cleaned_data.get('email')

            # Build query
            transactions = Transaction.objects.all().select_related('client', 'dealer', 'location', 'pos')

            if dealer:
                transactions = transactions.filter(dealer=dealer)
            if client:
                transactions = transactions.filter(client=client)
            if location:
                transactions = transactions.filter(location=location)
            if start_date:
                transactions = transactions.filter(created_at__date__gte=start_date)
            if end_date:
                transactions = transactions.filter(created_at__date__lte=end_date)

            # Generate Excel report
            from openpyxl import Workbook
            from django.core.mail import EmailMessage
            import io
            from django.utils import timezone

            wb = Workbook()
            ws = wb.active
            ws.title = "Transaction Report"

            # Add headers
            headers = ['ID', 'Transaction Type', 'Amount', 'Client', 'Dealer', 'Location', 'POS', 'Category', 'Description', 'Created At']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            # Add data
            for row_num, transaction in enumerate(transactions, 2):
                ws.cell(row=row_num, column=1, value=transaction.id)
                ws.cell(row=row_num, column=2, value=transaction.transaction_type)
                ws.cell(row=row_num, column=3, value=float(transaction.amount))
                ws.cell(row=row_num, column=4, value=str(transaction.client))
                ws.cell(row=row_num, column=5, value=str(transaction.dealer))
                ws.cell(row=row_num, column=6, value=str(transaction.location))
                ws.cell(row=row_num, column=7, value=str(transaction.pos))
                ws.cell(row=row_num, column=8, value=transaction.category)
                ws.cell(row=row_num, column=9, value=transaction.description)
                ws.cell(row=row_num, column=10, value=transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'))

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Send email
            subject = 'Transaction Report'
            message = f'Please find attached the transaction report generated on {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}.'
            email_msg = EmailMessage(subject, message, to=[email])
            email_msg.attach('transaction_report.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email_msg.send()

            messages.success(request, f'Report sent successfully to {email}!')
            return redirect('reports')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ReportForm()

    return render(request, 'admin/reports.html', {'form': form})

def admin_reset_password(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')
    
    username = request.session.get('admin_username')
    try:
        admin = Admin.objects.get(username=username)
        if request.method == 'POST':
            old_password = request.POST.get('old_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if not check_password(old_password, admin.password):
                 messages.error(request, 'Incorrect old password')
            elif new_password != confirm_password:
                messages.error(request, 'New passwords do not match')
            else:
                admin.password = make_password(new_password)
                admin.save()
                messages.success(request, 'Password changed successfully')
                return redirect('settings')
        
        context = {
            'username': username,
        }
        return render(request, 'admin/reset_password.html', context)
    except Admin.DoesNotExist:
        return redirect('admin_login')
