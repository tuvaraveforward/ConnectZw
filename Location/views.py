from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from .models import Location, User
from Admin.models import Transaction
from .forms import LocationReportForm, LocationFeedbackForm, LocationUserForm, LocationPosForm
from Pos.models import Pos
from Admin.utils import send_account_notification

# Create your views here.
@csrf_exempt
def location_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            user = Location.objects.get(username=username, is_active=True)
            if user.check_password(password):
                request.session.flush()
                request.session['location_logged_in'] = True
                request.session['location_id'] = user.id
                request.session['location_username'] = user.username
                return redirect('location_dashboard')  # Assuming you have a dashboard view
            else:
                messages.error(request, 'Invalid username or password.')
        except Location.DoesNotExist:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'Location/location_login.html')

def location_dashboard(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')

    location_id = request.session.get('location_id')
    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        request.session.flush()
        return redirect('location_login')

    return render(request, 'Location/location_dashboard.html', {'location': location})

def location_sales(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')

    location_id = request.session.get('location_id')
    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        request.session.flush()
        return redirect('location_login')

    from django.db.models import Q
    transactions = Transaction.objects.filter(
        Q(location=location) | Q(redemption_pos__location=location)
    ).select_related('client', 'pos', 'pos__user', 'redemption_pos', 'redemption_pos__user').order_by('-timestamp')
    
    return render(request, 'Location/location_sales.html', {'location': location, 'transactions': transactions})

def location_reports(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')

    location_id = request.session.get('location_id')
    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        request.session.flush()
        return redirect('location_login')

    if request.method == 'POST':
        form = LocationReportForm(request.POST)
        if form.is_valid():
            # Get filter parameters
            client = form.cleaned_data.get('client')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            email = form.cleaned_data.get('email')

            # Build query for the location
            transactions = Transaction.objects.filter(location=location).select_related('client', 'dealer', 'location', 'pos')

            if client:
                transactions = transactions.filter(client=client)
            if start_date:
                transactions = transactions.filter(created_at__date__gte=start_date)
            if end_date:
                transactions = transactions.filter(created_at__date__lte=end_date)

            # Generate Excel report
            from openpyxl import Workbook
            from django.core.mail import EmailMessage
            import io
            from django.utils import timezone

            wb = Workbook()
            ws = wb.active
            ws.title = "Transaction Report"

            # Add headers
            headers = ['ID', 'Transaction Type', 'Amount', 'Client', 'Dealer', 'Location', 'POS', 'Category', 'Description', 'Created At']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            # Add data
            for row_num, transaction in enumerate(transactions, 2):
                ws.cell(row=row_num, column=1, value=transaction.id)
                ws.cell(row=row_num, column=2, value=transaction.transaction_type)
                ws.cell(row=row_num, column=3, value=float(transaction.amount))
                ws.cell(row=row_num, column=4, value=str(transaction.client))
                ws.cell(row=row_num, column=5, value=str(transaction.dealer))
                ws.cell(row=row_num, column=6, value=str(transaction.location))
                ws.cell(row=row_num, column=7, value=str(transaction.pos))
                ws.cell(row=row_num, column=8, value=transaction.category)
                ws.cell(row=row_num, column=9, value=transaction.description)
                ws.cell(row=row_num, column=10, value=transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'))

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Send email
            subject = 'Transaction Report'
            message = f'Please find attached the transaction report for location "{location.location_name}" generated on {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}.'
            email_msg = EmailMessage(subject, message, to=[email])
            email_msg.attach('transaction_report.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email_msg.send()

            messages.success(request, f'Report sent successfully to {email}!')
            return redirect('location_reports')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = LocationReportForm()

    context = {
        'location': location,
        'form': form,
    }
    return render(request, 'Location/location_reports.html', context)

def location_feedback(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')

    location_id = request.session.get('location_id')
    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        request.session.flush()
        return redirect('location_login')

    if request.method == 'POST':
        form = LocationFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.location = location
            feedback.save()
            messages.success(request, 'Your feedback has been submitted successfully!')
            return redirect('location_feedback')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = LocationFeedbackForm()

    context = {
        'location': location,
        'form': form,
    }
    return render(request, 'Location/location_feedback.html', context)

def location_settings(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')

    location_id = request.session.get('location_id')
    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        request.session.flush()
        return redirect('location_login')

    available_pos = Location.objects.get(id=location_id).pos_devices.filter(user__isnull=True)
    return render(request, 'Location/location_settings.html', {
        'location': location,
        'users': location.users.all(),
        'pos_devices': location.pos_devices.all(),
        'available_pos': available_pos,
        'user_form': LocationUserForm(),
        'pos_form': LocationPosForm(location=location)
    })

def add_location_user(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')
    
    location_id = request.session.get('location_id')
    location = Location.objects.get(id=location_id)

    if request.method == 'POST':
        form = LocationUserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.location = location
            user.save()
            
            # Send email to the new User
            send_account_notification(user, form.cleaned_data['password'], 'user', request)
            
            messages.success(request, f'User {user.username} created successfully!')
        else:
             for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")
    
    return redirect('location_settings')

def add_location_pos(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')
        
    location_id = request.session.get('location_id')
    location = Location.objects.get(id=location_id)

    if request.method == 'POST':
        form = LocationPosForm(request.POST, location=location)
        if form.is_valid():
            pos = form.save(commit=False)
            pos.location = location
            pos.save()
            
            # Send email to Location (or assigned User if in future we modify this)
            send_account_notification(pos, pos.pos_code, 'pos', request)
            
            messages.success(request, f'POS {pos.pos_name} created successfully!')
        else:
             for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")

    return redirect('location_settings')

def assign_location_pos(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')
        
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        pos_id = request.POST.get('pos_id')
        
        try:
            user = User.objects.get(id=user_id)
            pos = Pos.objects.get(id=pos_id)
            
            # Verify both belong to current location for security
            location_id = request.session.get('location_id')
            if user.location.id == location_id and pos.location.id == location_id:
                pos.user = user
                pos.save()
                messages.success(request, f'POS {pos.pos_name} assigned to {user.firstname} {user.lastname}.')
            else:
                messages.error(request, 'Invalid User or POS for this location.')
                
        except (User.DoesNotExist, Pos.DoesNotExist):
            messages.error(request, 'User or POS not found.')
            
    return redirect('location_settings')

def location_about(request):
    if not request.session.get('location_logged_in'):
        return redirect('location_login')

    location_id = request.session.get('location_id')
    try:
        location = Location.objects.get(id=location_id)
    except Location.DoesNotExist:
        request.session.flush()
        return redirect('location_login')

    return render(request, 'Location/location_about.html', {'location': location})
