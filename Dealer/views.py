from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from Location.models import Location
from Pos.models import Pos
from Admin.models import Transaction, Product
from .models import Dealer
from .forms import DealerReportForm, ProductForm

from django.db.models import Sum

# Create your views here.
def dealer_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        try:
            dealer = Dealer.objects.get(username=username)
            if dealer.check_password(password):
                request.session['dealer_logged_in'] = True
                request.session['dealer_id'] = dealer.id
                request.session['dealer_username'] = dealer.username
                return redirect('dealer_dashboard')
            else:
                messages.error(request, 'Invalid username or password.')
        except Dealer.DoesNotExist:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'Dealer/dealer_login.html')

def dealer_dashboard(request):
    if not request.session.get('dealer_logged_in'):
        return redirect('dealer_login')
    
    dealer_id = request.session.get('dealer_id')
    try:
        dealer = Dealer.objects.get(id=dealer_id)
    except Dealer.DoesNotExist:
        return redirect('dealer_login')
    
    # Calculate sales by province for the dealer's category
    province_sales_data = Transaction.objects.filter(
        category=dealer.category, 
        transaction_type='purchase'
    ).values('location__province').annotate(
        total_sales=Sum('amount')
    ).order_by('location__province')

    # Convert query result to a dictionary for easy lookup
    sales_dict = {item['location__province']: float(item['total_sales']) for item in province_sales_data if item['location__province']}

    # Standard Zimbabwean Provinces as requested
    target_provinces = [
        'Harare', 'Bulawayo', 'Masvingo', 'Mat-North', 'Mat-South', 
        'Manicaland', 'Mash East', 'Mash West', 'Mash Central'
    ]

    province_labels = []
    province_sales = []

    for province in target_provinces:
        province_labels.append(province)
        # Real Data: Use actual sales or 0
        province_sales.append(sales_dict.get(province, 0))
    
    # Add any "Unknown" or other provinces not in the target list at the end
    # for province, sales in sales_dict.items():
    #     if province not in target_provinces:
    #         province_labels.append(province)
    #         province_sales.append(sales)

    import json
    context = {
        'dealer': dealer,
        'province_labels': json.dumps(province_labels),
        'province_sales': json.dumps(province_sales),
    }
    # --- Product sales (by product name) for the dealer's category ---
    # --- Product sales (by product name) for the dealer's category ---
    # We need:
    # 1. Overall product sales (for initial view / "All" provinces)
    # 2. Product sales broken down by province (for drill-down)
    
    import re
    import json
    
    # helper to parse product name from transaction description
    pattern = re.compile(r'Purchase:\s*(\d+)\s*x\s*(.+)', re.IGNORECASE)

    def get_product_sales_from_transactions(txs):
        sales_map = {}
        for t in txs:
            desc = (t.description or '').strip()
            m = pattern.match(desc)
            if m:
                pname = m.group(2).strip()
            else:
                pname = desc or 'Unknown'
            
            try:
                amt = float(t.amount)
            except Exception:
                amt = 0.0
            
            sales_map[pname] = sales_map.get(pname, 0.0) + amt
        return sales_map

    # 1. Overall Product Sales
    all_purchases = Transaction.objects.filter(category=dealer.category, transaction_type='purchase')
    overall_product_sales = get_product_sales_from_transactions(all_purchases)

    # 2. Product Sales by Province
    # Structure: {'Harare': {'Petrol': 100, 'Diesel': 50}, 'Bulawayo': ...}
    sales_by_province_data = {}
    
    # We can optimize by fetching all and iterating once, or filtering per province.
    # Given the scale likely isn't huge yet, iterating through the already fetched queryset (or similar) is fine.
    # We already have `province_sales_data` which is aggregated, but we need line items.
    
    # Let's iterate through `all_purchases` (which we already have) and bucket them by province.
    # Note: `all_purchases` needs to include location info.
    all_purchases = all_purchases.select_related('location')

    for t in all_purchases:
        if not t.location or not t.location.province:
            prov = 'Unknown'
        else:
            prov = t.location.province
        
        if prov not in sales_by_province_data:
            sales_by_province_data[prov] = {}
        
        # Parse product
        desc = (t.description or '').strip()
        m = pattern.match(desc)
        if m:
            pname = m.group(2).strip()
        else:
            pname = desc or 'Unknown'
        
        try:
            amt = float(t.amount)
        except Exception:
            amt = 0.0
            
        sales_by_province_data[prov][pname] = sales_by_province_data[prov].get(pname, 0.0) + amt

    # Prepare Context Data
    
    # Initial Product Chart (Overall)
    product_labels = list(overall_product_sales.keys())
    product_values = list(overall_product_sales.values())

    context['product_labels'] = json.dumps(product_labels)
    context['product_sales'] = json.dumps(product_values)
    context['sales_by_province_json'] = json.dumps(sales_by_province_data) # New context variable for JS

    return render(request, 'Dealer/dealer_dashboard.html', context)

def dealer_locations(request):
    if not request.session.get('dealer_logged_in'):
        return redirect('dealer_login')
    
    dealer_id = request.session.get('dealer_id')
    try:
        dealer = Dealer.objects.get(id=dealer_id)
    except Dealer.DoesNotExist:
        return redirect('dealer_login')
    
    # Get locations that match the dealer's category
    locations = Location.objects.filter(category=dealer.category, is_active=True)
    
    # Prepare data for the table
    locations_data = []
    for location in locations:
        pos_devices = Pos.objects.filter(location=location, is_active=True)
        locations_data.append({
            'location': location,
            'pos_count': pos_devices.count(),
            'pos_list': pos_devices,
        })
    
    context = {
        'dealer': dealer,
        'locations_data': locations_data,
    }
    return render(request, 'Dealer/dealer_locations.html', context)

def dealer_sales(request):
    if not request.session.get('dealer_logged_in'):
        return redirect('dealer_login')

    dealer_id = request.session.get('dealer_id')
    try:
        dealer = Dealer.objects.get(id=dealer_id)
    except Dealer.DoesNotExist:
        return redirect('dealer_login')

    # Get all transactions for the dealer's category
    transactions = Transaction.objects.filter(category=dealer.category).order_by('-timestamp')

    context = {
        'dealer': dealer,
        'transactions': transactions,
    }
    return render(request, 'Dealer/dealer_sales.html', context)

def dealer_reports(request):
    if not request.session.get('dealer_logged_in'):
        return redirect('dealer_login')

    dealer_id = request.session.get('dealer_id')
    try:
        dealer = Dealer.objects.get(id=dealer_id)
    except Dealer.DoesNotExist:
        return redirect('dealer_login')

    if request.method == 'POST':
        form = DealerReportForm(request.POST, category=dealer.category)
        if form.is_valid():
            # Get filter parameters
            client = form.cleaned_data.get('client')
            location = form.cleaned_data.get('location')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            email = form.cleaned_data.get('email')

            # Build query for dealer's category
            transactions = Transaction.objects.filter(category=dealer.category).select_related('client', 'dealer', 'location', 'pos')

            if client:
                transactions = transactions.filter(client=client)
            if location:
                transactions = transactions.filter(location=location)
            if start_date:
                transactions = transactions.filter(created_at__date__gte=start_date)
            if end_date:
                transactions = transactions.filter(created_at__date__lte=end_date)

            # Generate Excel report
            from openpyxl import Workbook
            from django.core.mail import EmailMessage
            import io
            from django.utils import timezone

            wb = Workbook()
            ws = wb.active
            ws.title = "Transaction Report"

            # Add headers
            headers = ['ID', 'Transaction Type', 'Amount', 'Client', 'Dealer', 'Location', 'POS', 'Category', 'Description', 'Created At']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            # Add data
            for row_num, transaction in enumerate(transactions, 2):
                ws.cell(row=row_num, column=1, value=transaction.id)
                ws.cell(row=row_num, column=2, value=transaction.transaction_type)
                ws.cell(row=row_num, column=3, value=float(transaction.amount))
                ws.cell(row=row_num, column=4, value=str(transaction.client))
                ws.cell(row=row_num, column=5, value=str(transaction.dealer))
                ws.cell(row=row_num, column=6, value=str(transaction.location))
                ws.cell(row=row_num, column=7, value=str(transaction.pos))
                ws.cell(row=row_num, column=8, value=transaction.category)
                ws.cell(row=row_num, column=9, value=transaction.description)
                ws.cell(row=row_num, column=10, value=transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'))

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Send email
            subject = 'Transaction Report'
            message = f'Please find attached the transaction report for category "{dealer.category}" generated on {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}.'
            email_msg = EmailMessage(subject, message, to=[email])
            email_msg.attach('transaction_report.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email_msg.send()

            messages.success(request, f'Report sent successfully to {email}!')
            return redirect('dealer_reports')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = DealerReportForm(category=dealer.category)

    context = {
        'dealer': dealer,
        'form': form,
    }
    return render(request, 'Dealer/dealer_reports.html', context)

def dealer_settings(request):
    if not request.session.get('dealer_logged_in'):
        return redirect('dealer_login')

    dealer_id = request.session.get('dealer_id')
    try:
        dealer = Dealer.objects.get(id=dealer_id)
    except Dealer.DoesNotExist:
        return redirect('dealer_login')

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.dealer = dealer
            product.category = dealer.category
            product.save()
            messages.success(request, 'Product added successfully!')
            return redirect('dealer_settings')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ProductForm()

    # Get all products for the dealer
    products = Product.objects.filter(dealer=dealer)

    context = {
        'dealer': dealer,
        'form': form,
        'products': products,
    }
    return render(request, 'Dealer/dealer_settings.html', context)


def dealer_product_edit(request, product_id):
    if not request.session.get('dealer_logged_in'):
        return redirect('dealer_login')

    dealer_id = request.session.get('dealer_id')
    try:
        dealer = Dealer.objects.get(id=dealer_id)
    except Dealer.DoesNotExist:
        return redirect('dealer_login')

    try:
        product = Product.objects.get(id=product_id, dealer=dealer)
    except Product.DoesNotExist:
        messages.error(request, 'Product not found')
        return redirect('dealer_settings')

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product updated successfully')
            return redirect('dealer_settings')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ProductForm(instance=product)

    context = {
        'dealer': dealer,
        'form': form,
        'product': product,
    }
    return render(request, 'Dealer/product_edit.html', context)


def dealer_product_delete(request, product_id):
    if not request.session.get('dealer_logged_in'):
        return redirect('dealer_login')

    dealer_id = request.session.get('dealer_id')
    try:
        dealer = Dealer.objects.get(id=dealer_id)
    except Dealer.DoesNotExist:
        return redirect('dealer_login')

    try:
        product = Product.objects.get(id=product_id, dealer=dealer)
    except Product.DoesNotExist:
        messages.error(request, 'Product not found')
        return redirect('dealer_settings')

    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully')
        return redirect('dealer_settings')

    context = {
        'dealer': dealer,
        'product': product,
    }
    return render(request, 'Dealer/product_confirm_delete.html', context)

def dealer_reset_password(request):
    if not request.session.get('dealer_logged_in'):
        return redirect('dealer_login')
    
    dealer_id = request.session.get('dealer_id')
    try:
        dealer = Dealer.objects.get(id=dealer_id)
        if request.method == 'POST':
            old_password = request.POST.get('old_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            # Using check_password directly as seen in login view, assuming Dealer has this method or uses make_password internally
            # However, standard practice if using check_password/make_password from imports
            if not dealer.check_password(old_password):
                 messages.error(request, 'Incorrect old password')
            elif new_password != confirm_password:
                messages.error(request, 'New passwords do not match')
            else:
                # If Dealer model has set_password method use it, else use make_password
                # Given login uses dealer.check_password(password), it likely mimics User model or has helpers.
                # To be safe and consistent with Client implementation which uses make_password manually:
                dealer.password = make_password(new_password)
                dealer.save()
                messages.success(request, 'Password changed successfully')
                return redirect('dealer_settings')
        
        context = {
            'dealer': dealer,
        }
        return render(request, 'Dealer/reset_password.html', context)
    except Dealer.DoesNotExist:
        return redirect('dealer_login')
